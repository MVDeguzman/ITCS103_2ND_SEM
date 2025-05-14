import tkinter as tk
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from tkinter import Image, PhotoImage

window = tk.Tk()
window.title("score tracker")
window.geometry("800x400")
# window.resizable(False, False)

def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "UserData"
    ws.append(["Name", "Score", "Remarks"])
    wb.save("userdata.xlsx")



def validate_inputs():
    name = name_entry.get()
    score= score_entry.get()

    if not name or not score:
        messagebox.showerror(title="Input Error",message= "All fields are required!")
        return False
    
def name_is_int():
    name = name_entry.get()
    if re.fullmatch(r'\d+', name): 
        messagebox.showerror(title="Input Error", message="Name must not be a number!")
        return True  
    return False
    
def score_failed():
    from openpyxl import load_workbook

def compute_average():
    wb = load_workbook("userdata.xlsx")
    ws = wb["UserData"]

    total = 0
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        score = row[1]  # Score is in the second column
        if isinstance(score, (int, float)):
            total += score
            count += 1

    if count > 0:
        average = total / count
        print(f"Average score: {average:.2f}")
    else:
        print("No scores found.")




def save_to_excel():

    if validate_inputs():
        return True
    
    if name_is_int():
        return True

    name = name_entry.get()
    score= int(score_entry.get())

    wb = load_workbook("userdata.xlsx")
    if "UserData" not in wb.sheetnames:
        create_excel_file()
    else:
        ws = wb["UserData"]
    
    if score >74 and score <=100:
        ws.append([name, score,"Pass"])
        messagebox.showinfo(title="Pass",message= "You have passed the exam!")

    elif score >0 and score <75:
        ws.append([name, score,"Fail"])
        messagebox.showinfo(title="Fail",message= "You have failed the exam!")

    elif score<0:
        messagebox.showerror(title="Input Error",message= "Score must be greater than 0!")
        return False
    
    elif score>100:
        messagebox.showerror(title="Input Error",message= "Score must be less than 100!")
        return False
    
    else:
        ws.append([name, score,"Invalid"])
        
    
    wb.save("userdata.xlsx")




  # Optional formatting
    messagebox.showinfo(title="Success",message= "Data saved successfully!")

    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)

    name_entry.insert(0, "Name")
    score_entry.insert(0, "Score")
    
def show_data():

    wb = load_workbook("userdata.xlsx")
    ws = wb["UserData"]

    data_window = tk.Toplevel(window)
    data_window.title("Stored User Data")
    data_window.geometry("400x400")



    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = tk.Label(data_window, text=value, borderwidth=1, relief="solid", padx=6, pady=3)
            label.grid(row=i, column=j)

# ========== Tkinter UI ==========
img = tk.PhotoImage(file = r'C:\Users\Mark Vincent\desktop\ITCCS103 2nd sem\quiz2_itcs103\812553.png')
image = Label(window, image=img)
image.place(x=0, y=0, relwidth=1, relheight=1)

mainframe= tk.Frame(window, bg= "gray25")
mainframe.place( anchor= "center", relx= 0.5, rely= 0.5,)

# img = tk.PhotoImage(file = r'C:\Users\Mark Vincent\desktop\ITCCS103 2nd sem\quiz2_itcs103\812553.png')
# Label(mainframe, image=img).place(x=0, y=0, relwidth=1, relheight=1)

user_frame= tk.LabelFrame(mainframe,bg= "gray30", )
user_frame.pack(pady=20, padx=20, fill= "both", expand= True)

title_frame= tk.Frame(user_frame, bg= "gray30")
title_frame.pack()

user= tk.Label(title_frame, text= "User Information", font= ("Arial", 20, "bold"),bg= "gray30", fg= "white")
user.pack()

input_frame= tk.Frame(user_frame, bg= "gray30")
input_frame.pack(fill= "both", expand= True)

name_label= tk.Label(input_frame, text= "Name:",font= ("Arial", 15, "bold"), fg= "white",bg= "gray30")
name_label.grid(row= 0, column= 0, padx= 10, pady= 10, sticky= "w")

name_entry= tk.Entry(input_frame, width= 30)
name_entry.grid(row= 0, column= 1, padx= 10, pady= 10)

score= tk.Label(input_frame, text= "Score Information:", font= ("Arial", 15, "bold" ), fg= "white",bg = "gray30")
score.grid(row= 1, column= 0, padx= 10, pady= 10, sticky= "w")

score_entry= tk.Entry(input_frame, width= 30)
score_entry.grid(row= 1, column= 1, padx= 10, pady= 10)

enter_button= tk.Button(input_frame, text= "Enter", font= ("Arial", 15), bg= "yellow2", command= save_to_excel)
enter_button.grid(row= 2, column= 0, padx= 10, pady= 10)

show_data_button= tk.Button(input_frame, text= "Show Data", font= ("Arial", 15), bg= "dodger blue", command= show_data)
show_data_button.grid(row= 2, column= 1, padx= 10, pady= 10)


window.mainloop()
